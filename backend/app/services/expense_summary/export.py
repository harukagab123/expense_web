from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from app.schemas.summary import ExpenseSummaryResponse


BLACK = "000000"
WHITE = "FFFFFF"
CURRENCY_FORMAT = '$#,##0.00;($#,##0.00);-'


@dataclass(frozen=True)
class ExportedWorkbook:
    content: bytes
    filename: str


def _safe_filename(summary: ExpenseSummaryResponse) -> str:
    if summary.period.mode == "TAX_YEAR" and summary.period.tax_year:
        suffix = str(summary.period.tax_year)
    else:
        suffix = f"{summary.period.start_date.isoformat()}-to-{summary.period.end_date.isoformat()}"
    return f"expense-summary-{suffix}.xlsx"


def export_expense_summary(summary: ExpenseSummaryResponse) -> ExportedWorkbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.sheet_view.showGridLines = False
    thin = Side(style="thin", color=BLACK)
    medium = Side(style="medium", color=BLACK)
    double = Side(style="double", color=BLACK)
    black_fill = PatternFill("solid", fgColor=BLACK)

    sheet.merge_cells("A1:B1")
    sheet["A1"] = f"{summary.period.label} EXPENSE SUMMARY"
    sheet["A1"].fill = black_fill
    sheet["A1"].font = Font(bold=True, color=WHITE, size=18)
    sheet.row_dimensions[1].height = 34
    sheet.merge_cells("A2:B2")
    sheet["A2"] = f"Reporting period: {summary.period.start_date} through {summary.period.end_date}"
    sheet["A2"].font = Font(italic=True, color=BLACK)
    sheet["A2"].border = Border(bottom=thin)

    sheet["A4"] = "TOTAL INCLUDED EXPENSES"
    sheet["B4"] = float(summary.grand_total)
    for cell in sheet[4]:
        cell.font = Font(bold=True, color=BLACK, size=13)
        cell.border = Border(top=medium, bottom=double)
    sheet["B4"].number_format = CURRENCY_FORMAT

    row = 7
    for group in summary.groups:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.cell(row, 1, group.label)
        sheet.cell(row, 1).fill = black_fill
        sheet.cell(row, 1).font = Font(bold=True, color=WHITE)
        sheet.row_dimensions[row].height = 24
        row += 1
        sheet.cell(row, 1, "Subcategory")
        sheet.cell(row, 2, "Amount")
        for cell in sheet[row]:
            cell.font = Font(bold=True, color=BLACK)
            cell.border = Border(bottom=medium)
        row += 1
        for subcategory in group.subcategories:
            sheet.cell(row, 1, subcategory.label)
            sheet.cell(row, 2, float(subcategory.total))
            sheet.cell(row, 2).number_format = CURRENCY_FORMAT
            for cell in sheet[row]:
                cell.border = Border(bottom=thin)
            row += 1
        sheet.cell(row, 1, f"TOTAL {group.label}")
        sheet.cell(row, 2, float(group.total))
        for cell in sheet[row]:
            cell.font = Font(bold=True, color=BLACK)
            cell.border = Border(top=medium, bottom=double)
        sheet.cell(row, 2).number_format = CURRENCY_FORMAT
        row += 2

    sheet.cell(row, 1, "TOTAL INCLUDED EXPENSES")
    sheet.cell(row, 2, float(summary.grand_total))
    for cell in sheet[row]:
        cell.fill = black_fill
        cell.font = Font(bold=True, color=WHITE, size=13)
        cell.border = Border(top=medium, bottom=double)
    sheet.cell(row, 2).number_format = CURRENCY_FORMAT
    sheet.column_dimensions["A"].width = 44
    sheet.column_dimensions["B"].width = 20
    sheet.freeze_panes = "A3"

    detail = workbook.create_sheet("Transaction Detail")
    detail.sheet_view.showGridLines = False
    headers = (
        "Transaction ID",
        "Date",
        "Name",
        "Transaction Detail",
        "Institution",
        "Source File",
        "Type",
        "Direction",
        "Main Category",
        "Subcategory",
        "Amount",
        "Category Status",
        "Review Status",
    )
    detail.append(headers)
    for cell in detail[1]:
        cell.fill = black_fill
        cell.font = Font(bold=True, color=WHITE)
        cell.border = Border(bottom=medium)

    detail_total = 0.0
    for group in summary.groups:
        for subcategory in group.subcategories:
            for transaction in subcategory.transactions:
                amount = float(transaction.amount)
                detail_total += amount
                detail.append(
                    (
                        transaction.id,
                        transaction.transaction_date,
                        transaction.normalized_name or "",
                        transaction.transaction_detail,
                        transaction.institution,
                        transaction.source_file,
                        transaction.transaction_type,
                        transaction.direction,
                        transaction.main_category_label or group.label,
                        transaction.subcategory_label or subcategory.label,
                        amount,
                        transaction.category_status,
                        transaction.review_status,
                    )
                )
                row_number = detail.max_row
                detail.cell(row_number, 2).number_format = "yyyy-mm-dd"
                detail.cell(row_number, 11).number_format = CURRENCY_FORMAT
                for cell in detail[row_number]:
                    cell.border = Border(bottom=thin)

    total_row = detail.max_row + 2
    detail.cell(total_row, 10, "TOTAL INCLUDED EXPENSES")
    detail.cell(total_row, 11, detail_total)
    for cell in detail[total_row]:
        cell.font = Font(bold=True, color=BLACK)
        cell.border = Border(top=medium, bottom=double)
    detail.cell(total_row, 11).number_format = CURRENCY_FORMAT
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = f"A1:M{max(1, detail.max_row - 2)}"
    for column, width in {
        "A": 16,
        "B": 13,
        "C": 24,
        "D": 42,
        "E": 20,
        "F": 28,
        "G": 24,
        "H": 12,
        "I": 32,
        "J": 30,
        "K": 16,
        "L": 20,
        "M": 18,
    }.items():
        detail.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    return ExportedWorkbook(content=output.getvalue(), filename=_safe_filename(summary))


def inspect_expense_summary_workbook(content: bytes, *, summary_preview_path: Path | None = None) -> dict:
    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["Summary"]
    detail = workbook["Transaction Detail"]
    values = [[cell.value for cell in row] for row in sheet.iter_rows()]
    detail_values = [[cell.value for cell in row] for row in detail.iter_rows()]
    formulas = [
        [cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None for cell in row]
        for row in sheet.iter_rows()
    ]
    formula_errors = [
        value for row_values in values for value in row_values
        if isinstance(value, str) and value.startswith("#")
    ]
    styles = []
    for inspected_sheet in (sheet, detail):
        for row_cells in inspected_sheet.iter_rows():
            for cell in row_cells:
                for color in (cell.font.color, cell.fill.fgColor):
                    rgb = getattr(color, "rgb", None)
                    if rgb:
                        styles.append({"value": str(rgb)[-6:]})
    if summary_preview_path is not None:
        summary_preview_path.parent.mkdir(parents=True, exist_ok=True)
        summary_preview_path.write_text(
            "Workbook preview rendering is not part of the packaged runtime.",
            encoding="utf-8",
        )
    return {
        "sheets": workbook.sheetnames,
        "summaryValues": values,
        "detailValues": detail_values,
        "summaryFormulas": formulas,
        "formulaErrors": formula_errors,
        "computedStyles": "\n".join(json.dumps(style) for style in styles),
        "sheetInspection": "Summary, Transaction Detail",
    }
